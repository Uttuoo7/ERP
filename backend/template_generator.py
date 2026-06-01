import io
import logging

logger = logging.getLogger(__name__)

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl or pandas is not installed. Fallback generators will be used.")

def generate_excel_template(module: str) -> bytes:
    """Generates a downloadable Excel template with instructions and data sheets."""
    if not OPENPYXL_AVAILABLE:
        # Fallback to simple CSV if libraries are missing
        return b"Vendor Name,GSTIN,PAN,State\n"
        
    wb = Workbook()
    
    # Sheet 1: Instructions
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="4F46E5") # Indigo
    
    ws_instructions['A1'] = f"{module.capitalize()} Import Template Instructions"
    ws_instructions['A1'].font = Font(bold=True, size=16)
    
    instructions_data = [
        ("Column Name", "Mandatory?", "Format/Rules", "Example"),
    ]
    
    if module == "vendors":
        instructions_data.extend([
            ("Vendor Name", "Yes", "Text", "Acme Corp"),
            ("GSTIN", "No", "15 char standard GSTIN", "27AAAAA0000A1Z5"),
            ("PAN", "No", "10 char standard PAN", "AAAAA0000A"),
            ("State", "No", "Text", "Maharashtra"),
        ])
    elif module == "items":
         instructions_data.extend([
            ("SKU", "Yes", "Unique Text", "ITM-001"),
            ("Name", "Yes", "Text", "Widget A"),
            ("Unit Price", "Yes", "Numeric", "100.50"),
            ("Description", "No", "Text", "A standard widget"),
        ])       

    for row_idx, row_data in enumerate(instructions_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
                
    ws_instructions.column_dimensions['A'].width = 25
    ws_instructions.column_dimensions['B'].width = 15
    ws_instructions.column_dimensions['C'].width = 30
    ws_instructions.column_dimensions['D'].width = 25
    
    # Sheet 2: Data Entry
    ws_data = wb.create_sheet(title="Data Entry")
    
    headers = [row[0] for row in instructions_data[1:]]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="1E293B") # Slate 800
        ws_data.column_dimensions[get_column_letter(col_idx)].width = 20
        
    # Freeze top row
    ws_data.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
